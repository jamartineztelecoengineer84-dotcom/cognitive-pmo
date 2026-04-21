"""
Gestor Documental Inteligente — API REST
Endpoints CRUD + búsqueda + stats + actividad
Tabla: primitiva.documentacion_repositorio
"""

import hashlib
import os
import json
import logging
from datetime import datetime
from typing import Optional

from fastapi import APIRouter, UploadFile, File, Form, Query, HTTPException
from fastapi.responses import FileResponse

from database import get_pool

logger = logging.getLogger("doc_routes")
router = APIRouter(prefix="/api/documentos", tags=["documentos"])

STORAGE_ROOT = "/app/data/documentos"
ALLOWED_MIME = {
    "application/pdf", "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "text/csv", "text/plain", "text/markdown", "text/html",
    "image/png", "image/jpeg", "application/octet-stream",
}


def _extraer_texto(content: bytes, mime: str) -> str:
    """Extrae texto plano del documento."""
    try:
        if "pdf" in mime:
            from PyPDF2 import PdfReader
            from io import BytesIO
            reader = PdfReader(BytesIO(content))
            return "\n".join(p.extract_text() or "" for p in reader.pages)[:50000]
        elif "wordprocessingml" in mime:
            from docx import Document
            from io import BytesIO
            doc = Document(BytesIO(content))
            return "\n".join(p.text for p in doc.paragraphs)[:50000]
        elif "spreadsheetml" in mime:
            from openpyxl import load_workbook
            from io import BytesIO
            wb = load_workbook(BytesIO(content), data_only=True)
            lines = []
            for ws in wb.worksheets:
                for row in ws.iter_rows(max_row=100, values_only=True):
                    lines.append(" | ".join(str(c) if c is not None else "" for c in row))
            return "\n".join(lines)[:50000]
        elif "text/" in mime or "csv" in mime:
            return content.decode("utf-8", errors="replace")[:50000]
    except Exception as e:
        logger.warning(f"Text extraction failed: {e}")
    return ""


async def _log_actividad(pool, doc_id, accion, usuario="system", detalle=None, ip=None):
    async with pool.acquire() as conn:
        await conn.execute(
            "INSERT INTO primitiva.doc_actividad_log (doc_id, accion, usuario, detalle, ip_origen) VALUES ($1,$2,$3,$4,$5)",
            doc_id, accion, usuario, json.dumps(detalle) if detalle else None, ip
        )


# ─── 1. UPLOAD ───

@router.post("/upload")
async def upload_documento(
    file: UploadFile = File(...),
    titulo: str = Form(None),
    silo: str = Form("transversal"),
    categoria: str = Form("organizacion"),
    proyecto_id: Optional[int] = Form(None),
    usuario: str = Form("system"),
    confidencialidad: str = Form("interna"),
    departamento: str = Form(None),
):
    content = await file.read()
    if len(content) == 0:
        raise HTTPException(400, "Archivo vacío")

    sha256 = hashlib.sha256(content).hexdigest()
    mime = file.content_type or "application/octet-stream"
    titulo_final = titulo or file.filename

    pool = get_pool()
    if not pool:
        raise HTTPException(503, "DB no disponible")

    # Deduplicación
    async with pool.acquire() as conn:
        dup = await conn.fetchval(
            "SELECT id FROM primitiva.documentacion_repositorio WHERE hash_sha256=$1 AND eliminado=false", sha256
        )
        if dup:
            return {"error": "duplicado", "existing_id": dup, "message": f"Archivo idéntico ya existe (id={dup})"}

    # Guardar archivo físico
    year = datetime.now().strftime("%Y")
    safe_name = f"{sha256[:8]}_{file.filename}"
    rel_path = f"{silo}/{year}/{safe_name}"
    full_path = os.path.join(STORAGE_ROOT, rel_path)
    os.makedirs(os.path.dirname(full_path), exist_ok=True)
    with open(full_path, "wb") as f:
        f.write(content)

    # Extraer texto
    texto = _extraer_texto(content, mime)
    num_paginas = None
    if "pdf" in mime:
        try:
            from PyPDF2 import PdfReader
            from io import BytesIO
            num_paginas = len(PdfReader(BytesIO(content)).pages)
        except Exception:
            pass

    # INSERT
    async with pool.acquire() as conn:
        row = await conn.fetchrow("""
            INSERT INTO primitiva.documentacion_repositorio
            (titulo, tipo, silo, categoria, proyecto_id, autor, mime_type, nombre_archivo,
             tamanio_bytes, hash_sha256, ruta_fisica, texto_extraido, num_paginas,
             confidencialidad, departamento, estado_procesamiento)
            VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11, $12, $13, $14, $15, 'pendiente')
            RETURNING id, titulo, estado_procesamiento
        """,
            titulo_final, _tipo_from_mime(mime), silo, categoria, proyecto_id, usuario,
            mime, file.filename, len(content), sha256, rel_path, texto, num_paginas,
            confidencialidad, departamento,
        )

    await _log_actividad(pool, row["id"], "upload", usuario, {"filename": file.filename, "size": len(content)})
    return {"id": row["id"], "titulo": row["titulo"], "estado_procesamiento": "pendiente", "size": len(content)}


def _tipo_from_mime(mime: str) -> str:
    if "pdf" in mime: return "pdf"
    if "word" in mime: return "docx"
    if "spreadsheet" in mime or "excel" in mime: return "xlsx"
    if "csv" in mime: return "csv"
    if "image" in mime: return "imagen"
    if "text/" in mime: return "texto"
    return "otro"


# ─── 2. LISTAR ───

@router.get("")
async def listar_documentos(
    silo: Optional[str] = None, tipo: Optional[str] = None,
    q: Optional[str] = None, categoria: Optional[str] = None,
    confidencialidad: Optional[str] = None, departamento: Optional[str] = None,
    estado_ia: Optional[str] = None,
    page: int = 1, size: int = 20, sort: str = "fecha_creacion DESC",
):
    pool = get_pool()
    if not pool:
        raise HTTPException(503, "DB no disponible")

    where, params = ["eliminado=false"], []
    idx = 1
    if silo:
        where.append(f"silo=${idx}"); params.append(silo); idx += 1
    if tipo:
        where.append(f"tipo=${idx}"); params.append(tipo); idx += 1
    if categoria:
        where.append(f"categoria=${idx}"); params.append(categoria); idx += 1
    if confidencialidad:
        where.append(f"confidencialidad=${idx}"); params.append(confidencialidad); idx += 1
    if departamento:
        where.append(f"departamento=${idx}"); params.append(departamento); idx += 1
    if estado_ia:
        where.append(f"estado_procesamiento=${idx}"); params.append(estado_ia); idx += 1
    if q:
        where.append(f"tsv @@ plainto_tsquery('spanish', ${idx})"); params.append(q); idx += 1

    where_sql = " AND ".join(where)
    allowed_sorts = {"fecha_creacion DESC", "fecha_creacion ASC", "titulo ASC", "titulo DESC", "tamanio_bytes DESC"}
    sort_sql = sort if sort in allowed_sorts else "fecha_creacion DESC"

    async with pool.acquire() as conn:
        total = await conn.fetchval(f"SELECT COUNT(*) FROM primitiva.documentacion_repositorio WHERE {where_sql}", *params)
        params.append(size); params.append((page - 1) * size)
        rows = await conn.fetch(
            f"""SELECT id, titulo, tipo, silo, categoria, autor, fecha_creacion, tamanio_bytes,
                       mime_type, nombre_archivo, estado_procesamiento, resumen_ia,
                       clasificacion_ia, palabras_clave_ia, confianza_ia, confidencialidad,
                       departamento, version
                FROM primitiva.documentacion_repositorio
                WHERE {where_sql} ORDER BY {sort_sql} LIMIT ${idx} OFFSET ${idx+1}""",
            *params
        )

    return {"items": [dict(r) for r in rows], "total": total, "page": page, "size": size}


# ─── 3. DETALLE ───

@router.get("/stats")
async def stats_documentos():
    pool = get_pool()
    if not pool:
        raise HTTPException(503, "DB no disponible")
    async with pool.acquire() as conn:
        total = await conn.fetchval("SELECT COUNT(*) FROM primitiva.documentacion_repositorio WHERE eliminado=false")
        by_tipo = await conn.fetch("SELECT tipo, COUNT(*) as c FROM primitiva.documentacion_repositorio WHERE eliminado=false GROUP BY tipo ORDER BY c DESC")
        by_silo = await conn.fetch("SELECT silo, COUNT(*) as c FROM primitiva.documentacion_repositorio WHERE eliminado=false GROUP BY silo ORDER BY c DESC")
        by_estado = await conn.fetch("SELECT estado_procesamiento, COUNT(*) as c FROM primitiva.documentacion_repositorio WHERE eliminado=false GROUP BY estado_procesamiento")
        total_size = await conn.fetchval("SELECT COALESCE(SUM(tamanio_bytes),0) FROM primitiva.documentacion_repositorio WHERE eliminado=false")
        papelera = await conn.fetchval("SELECT COUNT(*) FROM primitiva.documentacion_repositorio WHERE eliminado=true")
    return {
        "total": total, "papelera": papelera,
        "tamanio_total_bytes": total_size,
        "por_tipo": {r["tipo"]: r["c"] for r in by_tipo},
        "por_silo": {r["silo"]: r["c"] for r in by_silo},
        "por_estado_ia": {r["estado_procesamiento"]: r["c"] for r in by_estado},
    }


@router.get("/papelera")
async def listar_papelera():
    pool = get_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            "SELECT id, titulo, tipo, silo, nombre_archivo, eliminado_por, fecha_eliminacion FROM primitiva.documentacion_repositorio WHERE eliminado=true ORDER BY fecha_eliminacion DESC"
        )
    return {"items": [dict(r) for r in rows]}


@router.get("/actividad-global")
async def actividad_global():
    pool = get_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch("""
            SELECT a.*, d.titulo as doc_titulo FROM primitiva.doc_actividad_log a
            LEFT JOIN primitiva.documentacion_repositorio d ON a.doc_id = d.id
            ORDER BY a.created_at DESC LIMIT 50
        """)
    return {"items": [dict(r) for r in rows]}


@router.get("/{doc_id}")
async def detalle_documento(doc_id: int):
    pool = get_pool()
    async with pool.acquire() as conn:
        row = await conn.fetchrow("SELECT * FROM primitiva.documentacion_repositorio WHERE id=$1", doc_id)
    if not row:
        raise HTTPException(404, "Documento no encontrado")
    return dict(row)


# ─── 4. DESCARGAR ───

@router.get("/{doc_id}/descargar")
async def descargar_documento(doc_id: int):
    pool = get_pool()
    async with pool.acquire() as conn:
        row = await conn.fetchrow("SELECT ruta_fisica, nombre_archivo, mime_type FROM primitiva.documentacion_repositorio WHERE id=$1 AND eliminado=false", doc_id)
    if not row or not row["ruta_fisica"]:
        raise HTTPException(404, "Archivo no encontrado")
    full_path = os.path.join(STORAGE_ROOT, row["ruta_fisica"])
    if not os.path.exists(full_path):
        raise HTTPException(404, "Archivo físico no encontrado")
    await _log_actividad(pool, doc_id, "download")
    return FileResponse(full_path, filename=row["nombre_archivo"], media_type=row["mime_type"])


# ─── 5. PREVIEW ───

@router.get("/{doc_id}/preview")
async def preview_documento(doc_id: int):
    pool = get_pool()
    async with pool.acquire() as conn:
        row = await conn.fetchrow("SELECT texto_extraido, resumen_ia, titulo FROM primitiva.documentacion_repositorio WHERE id=$1", doc_id)
    if not row:
        raise HTTPException(404, "Documento no encontrado")
    return {"titulo": row["titulo"], "texto": (row["texto_extraido"] or "")[:5000], "resumen_ia": row["resumen_ia"]}


# ─── 6. EDITAR ───

@router.put("/{doc_id}")
async def editar_documento(doc_id: int, body: dict):
    pool = get_pool()
    allowed = {"titulo", "descripcion", "tipo", "silo", "categoria", "confidencialidad", "departamento"}
    updates, params = [], []
    idx = 1
    for k, v in body.items():
        if k in allowed:
            updates.append(f"{k}=${idx}"); params.append(v); idx += 1
    if not updates:
        raise HTTPException(400, "Nada que actualizar")
    updates.append(f"fecha_actualizacion=${idx}"); params.append(datetime.now()); idx += 1
    params.append(doc_id)
    async with pool.acquire() as conn:
        await conn.execute(f"UPDATE primitiva.documentacion_repositorio SET {','.join(updates)} WHERE id=${idx}", *params)
    await _log_actividad(pool, doc_id, "edit", detalle={"campos": list(body.keys())})
    return {"status": "updated", "id": doc_id}


# ─── 7. DELETE (soft) ───

@router.delete("/{doc_id}")
async def eliminar_documento(doc_id: int, usuario: str = "system"):
    pool = get_pool()
    async with pool.acquire() as conn:
        await conn.execute(
            "UPDATE primitiva.documentacion_repositorio SET eliminado=true, fecha_eliminacion=NOW(), eliminado_por=$1 WHERE id=$2",
            usuario, doc_id
        )
    await _log_actividad(pool, doc_id, "delete", usuario)
    return {"status": "deleted", "id": doc_id}


# ─── 8. RESTAURAR ───

@router.post("/{doc_id}/restaurar")
async def restaurar_documento(doc_id: int):
    pool = get_pool()
    async with pool.acquire() as conn:
        await conn.execute(
            "UPDATE primitiva.documentacion_repositorio SET eliminado=false, fecha_eliminacion=NULL, eliminado_por=NULL WHERE id=$1", doc_id
        )
    await _log_actividad(pool, doc_id, "restore")
    return {"status": "restored", "id": doc_id}


# ─── 9. REPROCESAR IA ───

@router.post("/{doc_id}/reprocesar")
async def reprocesar_documento(doc_id: int):
    pool = get_pool()
    async with pool.acquire() as conn:
        await conn.execute(
            "UPDATE primitiva.documentacion_repositorio SET estado_procesamiento='pendiente', resumen_ia=NULL, clasificacion_ia=NULL, palabras_clave_ia=NULL WHERE id=$1", doc_id
        )
    await _log_actividad(pool, doc_id, "reprocesar")
    return {"status": "queued", "id": doc_id}


# ─── 11. ACTIVIDAD DOC ───

@router.get("/{doc_id}/actividad")
async def actividad_documento(doc_id: int):
    pool = get_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            "SELECT * FROM primitiva.doc_actividad_log WHERE doc_id=$1 ORDER BY created_at DESC LIMIT 50", doc_id
        )
    return {"items": [dict(r) for r in rows]}


# ─── 12. VERSIONES ───

# ─── Q&A IA ───

@router.post("/pregunta")
async def pregunta_ia(body: dict):
    from agent_doc import responder_pregunta
    query = body.get("query") or body.get("mensaje", "")
    scope = body.get("scope", "all")
    if not query:
        raise HTTPException(400, "query es obligatorio")
    return await responder_pregunta(query, scope)


# ─── CHAT IA ───

@router.post("/chat")
async def chat_ia(body: dict):
    from agent_doc import responder_pregunta
    mensaje = body.get("mensaje", "")
    scope = body.get("scope", "all")
    if not mensaje:
        raise HTTPException(400, "mensaje es obligatorio")
    return await responder_pregunta(mensaje, scope)


@router.get("/{doc_id}/versiones")
async def versiones_documento(doc_id: int):
    pool = get_pool()
    async with pool.acquire() as conn:
        row = await conn.fetchrow("SELECT id, version, version_anterior_id FROM primitiva.documentacion_repositorio WHERE id=$1", doc_id)
        if not row:
            raise HTTPException(404)
        chain = [dict(row)]
        prev_id = row["version_anterior_id"]
        while prev_id:
            prev = await conn.fetchrow("SELECT id, version, version_anterior_id, titulo, fecha_creacion FROM primitiva.documentacion_repositorio WHERE id=$1", prev_id)
            if not prev:
                break
            chain.append(dict(prev))
            prev_id = prev["version_anterior_id"]
    return {"versiones": chain}
